# pos/views/manualexcel_views.py

import pandas as pd
import re
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.http import HttpResponse
from rest_framework import status
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from pos.models.items import items, itemvariants
from pos.models.group_unit import ItemGroup, ItemUnit
from django.db import transaction
from collections import defaultdict


from ecommerce.permissions import IsSuperAdminOrBranchOrPagePermittedEmployee


# ─────────────────────────────────────────────────────────────────────────────
# MAIN CRUD VIEWS (with permission check)
# ─────────────────────────────────────────────────────────────────────────────

class ManualDownloadExcelTemplate(APIView):
    """Download Excel template for manual items import"""
    
    # ✅ CHANGE: IsAuthenticated → IsSuperAdminOrBranchOrPagePermittedEmployee
    permission_classes = [IsSuperAdminOrBranchOrPagePermittedEmployee]
    page_key = "/ExcelImportExport"  # ✅ ADD: Frontend route

    def get(self, request):
        # ✅ CHANGE: request.user.branch → get_effective_branch()
        branch = request.user.get_effective_branch()
        if not branch:
            return Response({
                "success": False,
                "error": "No branch linked to this user"
            }, status=status.HTTP_400_BAD_REQUEST)
            
        branch_type = branch.branch_type.lower()

        wb = Workbook()
        ws_data = wb.active
        ws_data.title = "Item Data"

        # ===== STYLES =====
        header_fill = PatternFill(start_color="1F7A4A", end_color="1F7A4A", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # ===== COLUMNS =====
        main_columns = [
            {"header": "ITEM_NAME*",            "width": 30},
            {"header": "BRAND_NAME",             "width": 25},
            {"header": "CATEGORY_NAME*",          "width": 25},
            {"header": "SUB_CATEGORY_NAME",      "width": 25},
            {"header": "SUB_SUB_CATEGORY_NAME",  "width": 30},
            {"header": "GROUP_NAME",             "width": 20},
            {"header": "UNIT_NAME*",              "width": 15},
            {"header": "HSN_CODE*",               "width": 15},
            {"header": "TAX_SLAB",               "width": 15},
        ]

        branch_variant_fields = {
            "fashion":     ["VARIANT_SIZE", "VARIANT_COLOR"],
            "mart":        ["VARIANT_SIZE"],
            "electronics": ["VARIANT_SIZE", "VARIANT_COLOR", "SERIAL_NO", "WARRANTY_DATE"],
        }

        variant_columns = branch_variant_fields.get(branch_type, []) + [
            "PURCHASE_PRICE*", "SALES_PRICE*", "MRP*", "BARCODE", "OPENING_STOCK"
        ]

        all_headers = main_columns + [{"header": col, "width": 20} for col in variant_columns]

        # ===== WRITE HEADERS =====
        for col_idx, col in enumerate(all_headers, 1):
            cell = ws_data.cell(row=1, column=col_idx, value=col["header"])
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            ws_data.column_dimensions[get_column_letter(col_idx)].width = col["width"]

        # ===== DROPDOWNS — only GROUP, UNIT, TAX_SLAB =====
        dropdown_map = {
            "GROUP_NAME": sorted(list(ItemGroup.objects.filter(branch=branch).values_list('name', flat=True))),
            "UNIT_NAME":  sorted(list(ItemUnit.objects.filter(is_active=True).values_list('symbol', flat=True))),
            "TAX_SLAB":   ["5%", "12%", "18%", "28%", "Tax Free"],
        }

        hidden_start = len(all_headers) + 2
        hidden_cols = {}

        for i, (key, values) in enumerate(dropdown_map.items()):
            col_idx = hidden_start + i
            col_letter = get_column_letter(col_idx)
            hidden_cols[key] = col_letter

            for r, val in enumerate(values, start=2):
                ws_data.cell(row=r, column=col_idx, value=val)

            ws_data.column_dimensions[col_letter].hidden = True

        # ===== APPLY DROPDOWNS =====
        max_row = 500

        for col_idx, col in enumerate(all_headers, 1):
            col_letter = get_column_letter(col_idx)
            header = col["header"].replace("*", "")

            if header in hidden_cols:
                hidden_col = hidden_cols[header]
                length = len(dropdown_map[header])
                formula = f"${hidden_col}$2:${hidden_col}${length + 1}"
                dv = DataValidation(type="list", formula1=formula, allow_blank=True)
                ws_data.add_data_validation(dv)
                dv.add(f"{col_letter}2:{col_letter}{max_row}")

        # ===== CONDITIONAL FORMATTING =====
        header_names = [col["header"].replace("*", "") for col in all_headers]
        try:
            sales_col_letter = get_column_letter(header_names.index("SALES_PRICE") + 1)
            mrp_col_letter = get_column_letter(header_names.index("MRP") + 1)
            red_fill = PatternFill(start_color="FFC7CE", fill_type="solid")
            ws_data.conditional_formatting.add(
                f"{sales_col_letter}2:{sales_col_letter}{max_row}",
                CellIsRule(operator='greaterThan', formula=[f"${mrp_col_letter}2"], fill=red_fill)
            )
        except ValueError:
            pass

        # ===== RESPONSE =====
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="manual_item_template_{branch_type}.xlsx"'
        wb.save(response)
        return response


class ManualImportItemsFromExcel(APIView):
    """Import manual items from Excel"""
    
    # ✅ CHANGE: IsAuthenticated → IsSuperAdminOrBranchOrPagePermittedEmployee
    permission_classes = [IsSuperAdminOrBranchOrPagePermittedEmployee]
    page_key = "/ExcelImportExport"  # ✅ ADD: Frontend route

    def post(self, request):
        used_barcodes = set()

        excel_file = request.FILES.get('file')
        if not excel_file:
            return Response({"error": "No file uploaded"}, status=400)

        # ✅ CHANGE: request.user.branch → get_effective_branch()
        branch = request.user.get_effective_branch()
        if not branch:
            return Response({
                "success": False,
                "error": "No branch linked to this user"
            }, status=status.HTTP_400_BAD_REQUEST)
            
        branch_type = branch.branch_type.lower()

        variant_field_mapping = {
            "fashion":     {"size": "VARIANT_SIZE", "color": "VARIANT_COLOR"},
            "mart":        {"size": "VARIANT_SIZE"},
            "electronics": {
                "size":         "VARIANT_SIZE",
                "color":        "VARIANT_COLOR",
                "srno":         "SERIAL_NO",
                "warrantydate": "WARRANTY_DATE",
            },
        }

        # ===== HELPERS =====
        def is_empty(val):
            if val is None:
                return True
            if isinstance(val, float) and pd.isna(val):
                return True
            if str(val).strip() == "":
                return True
            return False

        def clean(val):
            if is_empty(val):
                return ""
            return str(val).strip()

        def to_float(val):
            try:
                if is_empty(val):
                    return None
                val = str(val).strip().replace(",", "")
                return float(val) if val else None
            except:
                return None

        def to_int(val):
            try:
                if is_empty(val):
                    return None
                val = str(val).strip()
                return int(float(val)) if val else None
            except:
                return None

        def normalize_tax(val):
            if is_empty(val):
                return ""

            if isinstance(val, (int, float)):
                num = float(val)
                if num == 0:
                    return "Tax Free"
                if 0 < num < 1:
                    num = round(num * 100, 2)
                valid_slabs = {5.0: "5%", 12.0: "12%", 18.0: "18%", 28.0: "28%"}
                return valid_slabs.get(num, "")

            cleaned = re.sub(r"\s+", "", str(val)).upper()

            if "TAXFREE" in cleaned or cleaned in ("NIL", "0%", "0"):
                return "Tax Free"

            match = re.search(r"(\d+(\.\d+)?)", cleaned)
            if not match:
                return ""

            num = float(match.group(1))
            valid_slabs = {5.0: "5%", 12.0: "12%", 18.0: "18%", 28.0: "28%"}
            return valid_slabs.get(num, "")

        def clean_numeric_text_cell(x):
            if pd.isna(x):
                return None
            if isinstance(x, float) and x.is_integer():
                return str(int(x))
            return str(x).strip()

        # ===== READ FILE =====
        try:
            df = pd.read_excel(
                excel_file,
                sheet_name=0,
                header=0,
                converters={
                    'BARCODE': clean_numeric_text_cell,
                    'HSN_CODE': clean_numeric_text_cell,
                    'HSN_CODE*': clean_numeric_text_cell,
                }
            )
        except Exception as e:
            return Response({"error": str(e)}, status=400)

        # ===== CLEAN HEADERS =====
        df.columns = [str(c).strip().replace('*', '').upper().replace(" ", "_") for c in df.columns]
        df = df.where(pd.notnull(df), None)

        if "ITEM_NAME" not in df.columns:
            return Response({
                "error": f"Invalid Excel format. Found columns: {list(df.columns)}"
            }, status=400)

        required_cols = ['ITEM_NAME', 'HSN_CODE', 'UNIT_NAME', 'CATEGORY_NAME',
                         'PURCHASE_PRICE', 'SALES_PRICE', 'MRP']
        missing = [c for c in required_cols if c not in df.columns]
        if missing:
            return Response({"error": f"Missing columns: {', '.join(missing)}"}, status=400)

        # ===== LOOKUPS =====
        group_map = {g.name: g.id for g in ItemGroup.objects.filter(branch=branch)}
        unit_map = {u.symbol: u.id for u in ItemUnit.objects.filter(is_active=True)}
        
        existing_barcodes = set(
            itemvariants.objects.filter(item__branch=branch)
            .exclude(barcode__isnull=True)
            .exclude(barcode="")
            .values_list("barcode", flat=True)
        )
        
        existing_item_names = set(
            name.strip().lower()
            for name in items.objects.filter(branch=branch).values_list("itemName", flat=True)
        )

        errors = []
        items_data = defaultdict(lambda: {
            "itemName":    "",
            "item_fields": {},
            "variants":    [],
            "has_error":   False,
        })

        last_item = None

        # ===== LOOP =====
        for idx, row in df.iterrows():
            row_num = idx + 2

            item_name = clean(row.get('ITEM_NAME'))

            if not item_name and not last_item:
                continue
            elif not item_name:
                item_name = last_item
            else:
                last_item = item_name

            if all(is_empty(row.get(c)) for c in ['ITEM_NAME', 'PURCHASE_PRICE', 'SALES_PRICE', 'MRP']):
                continue

            if items_data[item_name].get("has_error"):
                continue

            # ===== ITEM FIELDS =====
            if not items_data[item_name]["item_fields"]:
                val_category = clean(row.get('CATEGORY_NAME'))
                val_subcat = clean(row.get('SUB_CATEGORY_NAME'))
                val_subsub = clean(row.get('SUB_SUB_CATEGORY_NAME'))
                val_brand = clean(row.get('BRAND_NAME'))
                val_group = clean(row.get('GROUP_NAME'))
                val_unit = clean(row.get('UNIT_NAME'))
                val_hsn = clean(row.get('HSN_CODE'))

                item_has_error = False

                if not item_name:
                    errors.append(f"Row {row_num}: ITEM_NAME is required")
                    item_has_error = True
                elif item_name.strip().lower() in existing_item_names:
                    errors.append(f"Row {row_num}: Item '{item_name}' already exists in this branch")
                    item_has_error = True

                if not val_category:
                    errors.append(f"Row {row_num}: CATEGORY_NAME is required")
                    item_has_error = True

                if val_subcat and not val_category:
                    errors.append(f"Row {row_num}: SUB_CATEGORY_NAME provided but CATEGORY_NAME is empty")
                    item_has_error = True

                if val_subsub and not val_subcat:
                    errors.append(f"Row {row_num}: SUB_SUB_CATEGORY_NAME provided but SUB_CATEGORY_NAME is empty")
                    item_has_error = True

                if not val_unit:
                    errors.append(f"Row {row_num}: UNIT_NAME is required")
                    item_has_error = True
                elif not unit_map.get(val_unit):
                    errors.append(f"Row {row_num}: UNIT_NAME '{val_unit}' not found")
                    item_has_error = True

                if not val_hsn:
                    errors.append(f"Row {row_num}: HSN_CODE is required")
                    item_has_error = True

                val_tax = normalize_tax(row.get('TAX_SLAB'))
                if not val_tax:
                    errors.append(f"Row {row_num}: TAX_SLAB is required/invalid (allowed: 5%, 12%, 18%, 28%, Tax Free)")
                    item_has_error = True

                if item_has_error:
                    items_data[item_name]["has_error"] = True
                    continue

                items_data[item_name]["item_fields"] = {
                    "brand_name":     val_brand,
                    "category_name":  val_category,
                    "subcat_name":    val_subcat,
                    "subsub_name":    val_subsub,
                    "group":          group_map.get(val_group) if val_group else None,
                    "unit":           unit_map.get(val_unit),
                    "hsn":            val_hsn,
                    "tax":            val_tax,
                }

            # ===== PRICE =====
            purchase_price = to_float(row.get('PURCHASE_PRICE'))
            sales_price = to_float(row.get('SALES_PRICE'))
            mrp = to_float(row.get('MRP'))

            if purchase_price is None or sales_price is None or mrp is None:
                errors.append(f"Row {row_num}: PURCHASE_PRICE, SALES_PRICE, MRP are required")
                continue

            if purchase_price <= 0 or sales_price <= 0 or mrp <= 0:
                errors.append(f"Row {row_num}: Price must be greater than 0")

            if sales_price > mrp:
                errors.append(f"Row {row_num}: SALES_PRICE cannot be greater than MRP")

            # ===== BARCODE =====
            raw_barcode = row.get('BARCODE')
            if is_empty(raw_barcode):
                barcode = None
            else:
                barcode = str(raw_barcode).strip().upper()
                if barcode in ("", "NONE", "NAN", "NULL"):
                    barcode = None

            # ===== STOCK =====
            qty = to_int(row.get('OPENING_STOCK')) or 0

            # ===== TAX =====
            tax_str = normalize_tax(row.get('TAX_SLAB'))
            tax_rate = 0.0
            if tax_str and tax_str != "Tax Free":
                try:
                    tax_rate = float(tax_str.replace('%', ''))
                except:
                    tax_rate = 0.0

            basic = qty * purchase_price
            tax_amt = (basic * tax_rate) / 100
            net = basic + tax_amt

            variant = {
                "purchasePrice": purchase_price,
                "salesPrice":    sales_price,
                "mrp":           mrp,
                "barcode":       barcode,
                "opStock":       qty,
                "basicAmount":   basic,
                "taxAmount":     tax_amt,
                "netValue":      net,
            }

            # ===== VARIANT EXTRA FIELDS =====
            mapping = variant_field_mapping.get(branch_type, {})
            for key, col in mapping.items():
                val = clean(row.get(col))

                if key == "warrantydate" and val:
                    try:
                        val = pd.to_datetime(val)
                    except:
                        errors.append(f"Row {row_num}: Invalid WARRANTY_DATE")
                        val = None

                variant[key] = val if val else None

            # ===== DUPLICATE BARCODE =====
            if barcode:
                if barcode in used_barcodes:
                    errors.append(f"Row {row_num}: Duplicate barcode '{barcode}' (repeated in this file)")
                    continue
                if barcode in existing_barcodes:
                    errors.append(f"Row {row_num}: Barcode '{barcode}' already exists in this branch")
                    continue
                used_barcodes.add(barcode)

            items_data[item_name]["variants"].append(variant)
            items_data[item_name]["itemName"] = item_name

        if errors:
            return Response({"success": False, "errors": errors[:100]}, status=400)

        if not any(item["variants"] for item in items_data.values()):
            return Response({
                "success": False,
                "errors": ["Excel file is empty or no valid rows found"]
            }, status=400)

        created_items = 0
        created_variants = 0

        # ===== SAVE =====
        with transaction.atomic():
            for item_info in items_data.values():
                if not item_info["variants"]:
                    continue

                f = item_info["item_fields"]

                item_obj = items.objects.create(
                    itemName=item_info["itemName"],
                    branch=branch,
                    entry_type="manual",
                    brand=f["brand_name"],
                    category=f["category_name"],
                    subCategory=f["subcat_name"],
                    subSubCategory=f["subsub_name"],
                    group_id=f["group"],
                    unit_id=f["unit"],
                    hsnCode=f["hsn"],
                    taxSlab=f["tax"],
                )
                created_items += 1

                for v in item_info["variants"]:
                    itemvariants.objects.create(
                        item=item_obj,
                        purchasePrice=v["purchasePrice"],
                        salesPrice=v["salesPrice"],
                        mrp=v["mrp"],
                        barcode=v["barcode"],
                        opStock=v["opStock"],
                        basicAmount=v["basicAmount"],
                        taxAmount=v["taxAmount"],
                        netValue=v["netValue"],
                        current_stock=v["opStock"],
                        size=v.get("size"),
                        color=v.get("color"),
                        srno=v.get("srno"),
                        warrantydate=v.get("warrantydate"),
                    )
                    created_variants += 1

        return Response({
            "success": True,
            "items_created": created_items,
            "variants_created": created_variants,
        }, status=201)


class ManualExportItemsToExcel(APIView):
    """Export manual items to Excel"""
    
    # ✅ CHANGE: IsAuthenticated → IsSuperAdminOrBranchOrPagePermittedEmployee
    permission_classes = [IsSuperAdminOrBranchOrPagePermittedEmployee]
    page_key = "/ExcelImportExport"  # ✅ ADD: Frontend route

    def get(self, request):
        from openpyxl.utils import get_column_letter
        import pandas as pd

        # ✅ CHANGE: request.user.branch → get_effective_branch()
        branch = request.user.get_effective_branch()
        if not branch:
            return Response({
                "success": False,
                "error": "No branch linked to this user"
            }, status=status.HTTP_400_BAD_REQUEST)
            
        branch_type = branch.branch_type.lower()

        item_queryset = items.objects.filter(
            branch=branch, entry_type='manual'
        ).prefetch_related('variants')

        data = []
        for item in item_queryset:
            for variant in item.variants.all():
                row = {
                    "ITEM_NAME":             item.itemName,
                    "BRAND_NAME":            item.brand or "",
                    "CATEGORY_NAME":         item.category or "",
                    "SUB_CATEGORY_NAME":     item.subCategory or "",
                    "SUB_SUB_CATEGORY_NAME": item.subSubCategory or "",
                    "GROUP_NAME":            item.group.name if item.group else "",
                    "UNIT_NAME":             item.unit.symbol if item.unit else "",
                    "HSN_CODE":              item.hsnCode or "",
                    "TAX_SLAB":              item.taxSlab or "",
                    "VARIANT_SIZE":          variant.size or "",
                    "VARIANT_COLOR":         variant.color or "",
                    "SERIAL_NO":             variant.srno or "",
                    "WARRANTY_DATE":         variant.warrantydate.strftime('%Y-%m-%d') if variant.warrantydate else "",
                    "PURCHASE_PRICE":        variant.purchasePrice,
                    "SALES_PRICE":           variant.salesPrice,
                    "MRP":                   variant.mrp,
                    "BARCODE":               variant.barcode or "",
                    "OPENING_STOCK":         variant.opStock,
                }
                data.append(row)

        df = pd.DataFrame(data)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="manual_items_export.xlsx"'

        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Item Data")

            worksheet = writer.sheets["Item Data"]

            for col_idx, col in enumerate(df.columns, 1):
                max_length = len(str(col))

                for cell in df[col]:
                    try:
                        if cell:
                            max_length = max(max_length, len(str(cell)))
                    except:
                        pass

                if "BARCODE" in col:
                    width = 25
                elif "NAME" in col:
                    width = 30
                elif "PRICE" in col or "MRP" in col:
                    width = 18
                elif "DATE" in col:
                    width = 20
                else:
                    width = max_length + 3

                worksheet.column_dimensions[get_column_letter(col_idx)].width = width

        return response